import os
import sys
import datetime
import logging
import pandas as pd
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def clean_excel_to_csv(excel_path: str, output_csv_path: str):
    logging.info(f"Opening workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    selected_sheet_name = None
    
    # Locate active, visible sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state == 'visible':
            # Check cell value logic (e.g., cell B9 non-null)
            b9_val = ws['B9'].value
            if b9_val is not None:
                selected_sheet_name = sheet_name
                break
                
    if not selected_sheet_name:
        selected_sheet_name = wb.sheetnames[0]
        logging.warning(f"No sheet matching criteria; defaulting to {selected_sheet_name}")

    # Load sheet data into pandas dataframe
    df = pd.read_excel(excel_path, sheet_name=selected_sheet_name, header=None)
    
    # Process multi-row headers (e.g., rows 5-8 / indices 4-7)
    header_rows = df.iloc[4:8].fillna("").astype(str)
    
    clean_headers = []
    for col in header_rows.columns:
        combined = "".join(header_rows[col])
        for char in [" ", "(", ")", "#", ",", ".", "-", "+", "&", "\r", "\n"]:
            combined = combined.replace(char, "")
        clean_headers.append(combined.upper())
        
    # Apply cleaned headers to remaining data
    data_df = df.iloc[8:].copy()
    data_df.columns = clean_headers
    
    # Export sanitized CSV
    data_df.to_csv(output_csv_path, index=False)
    logging.info(f"Successfully converted data to {output_csv_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        logging.error("Missing input file argument.")
        sys.exit(1)
        
    input_file = sys.argv[1]
    output_file = os.path.splitext(input_file)[0] + ".csv"
    clean_excel_to_csv(input_file, output_file)
